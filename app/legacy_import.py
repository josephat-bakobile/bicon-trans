"""One-off import of the old per-month xlsx ledger into the SQLite models.

The old sheet laid out four fixed 4-column car blocks (MAKUANYO, MALIPO YA DENI,
MATUMIZI, MAELEZO) per day row. We reconstruct that history as:
  - one CollectionTransaction per date, with one line per car that had a
    MAKUANYO value that day (multiple cars collected "in one run" that day)
  - one ConsumptionEntry per car/day with a MATUMIZI value
  - one DebtPayment per car/day with a MALIPO YA DENI value
  - one opening Debt per car from the MADENI/INADAIWA row, if present
"""

import calendar
import datetime as dt

import openpyxl

from .extensions import db
from .models import Car, CollectionLine, CollectionTransaction, ConsumptionEntry, Debt, DebtPayment, ExpenseCategory

MONTH_NAMES = [
    "JANUARY", "FEBRUARY", "MARCH", "APRIL", "MAY", "JUNE",
    "JULY", "AUGUST", "SEPTEMBER", "OCTOBER", "NOVEMBER", "DECEMBER",
]


def _find_month_sheets(wb):
    sheets = []
    for name in wb.sheetnames:
        for i, month in enumerate(MONTH_NAMES, start=1):
            if name.upper().startswith(month):
                parts = name.split()
                try:
                    year = 2000 + int(parts[-1])
                except ValueError:
                    continue
                sheets.append((year, i, wb[name]))
                break
    return sheets


def _get_car_codes(ws):
    codes = []
    col = 2
    while ws.cell(row=2, column=col).value == "MAKUANYO":
        codes.append(ws.cell(row=1, column=col).value)
        col += 4
    return codes


def import_legacy(xlsx_path):
    if CollectionTransaction.query.count() > 0:
        return "skipped: collection transactions already exist"

    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=False)
    except FileNotFoundError:
        return "skipped: legacy file not found"

    month_sheets = _find_month_sheets(wb)
    if not month_sheets:
        return "skipped: no month sheets found"

    category = ExpenseCategory.query.filter_by(name="MATUMIZI").first()
    if category is None:
        category = ExpenseCategory(name="MATUMIZI")
        db.session.add(category)
        db.session.flush()

    imported_days = 0
    trans_counter = 0

    for year, month, ws in month_sheets:
        car_codes = _get_car_codes(ws)
        if not car_codes:
            continue
        cars = {}
        for code in car_codes:
            car = Car.query.filter_by(code=code).first()
            if car is None:
                car = Car(code=code)
                db.session.add(car)
                db.session.flush()
            cars[code] = car

        n_days = calendar.monthrange(year, month)[1]
        for day in range(1, n_days + 1):
            row = 2 + day
            cell_a = ws.cell(row=row, column=1).value
            if not isinstance(cell_a, (dt.date, dt.datetime)):
                continue
            entry_date = cell_a.date() if isinstance(cell_a, dt.datetime) else cell_a

            collection_lines = []
            for i, code in enumerate(car_codes):
                base = 2 + i * 4
                makuanyo = ws.cell(row=row, column=base).value
                malipo = ws.cell(row=row, column=base + 1).value
                matumizi = ws.cell(row=row, column=base + 2).value
                maelezo = ws.cell(row=row, column=base + 3).value
                maelezo = str(maelezo).strip() if maelezo else None
                car = cars[code]

                if isinstance(makuanyo, (int, float)) and makuanyo:
                    collection_lines.append((car, float(makuanyo), maelezo if not matumizi else None))
                if isinstance(matumizi, (int, float)) and matumizi:
                    db.session.add(
                        ConsumptionEntry(
                            date=entry_date, car_id=car.id, category_id=category.id,
                            amount=float(matumizi), description=maelezo,
                        )
                    )
                if isinstance(malipo, (int, float)) and malipo:
                    db.session.add(
                        DebtPayment(
                            date=entry_date, car_id=car.id, amount=float(malipo),
                            description=maelezo if not matumizi else None,
                        )
                    )

            if collection_lines:
                trans_counter += 1
                txn = CollectionTransaction(
                    date=entry_date, note="Imported from legacy sheet",
                    trans_no=f"TRX-{trans_counter:05d}",
                )
                db.session.add(txn)
                db.session.flush()
                for car, amount, note in collection_lines:
                    db.session.add(CollectionLine(transaction_id=txn.id, car_id=car.id, amount=amount, note=note))
                imported_days += 1

        # Opening debt balances from the MADENI/INADAIWA rows, if present.
        madeni_row = None
        for r in range(1, ws.max_row + 1):
            if ws.cell(row=r, column=1).value == "MADENI":
                madeni_row = r
                break
        if madeni_row:
            for i, code in enumerate(car_codes):
                r = madeni_row + 2 + i
                if ws.cell(row=r, column=1).value != code:
                    continue
                inadaiwa = ws.cell(row=r, column=2).value
                if isinstance(inadaiwa, (int, float)) and inadaiwa:
                    db.session.add(
                        Debt(
                            date=dt.date(year, month, 1),
                            car_id=cars[code].id,
                            amount=float(inadaiwa),
                            description=f"Salio la awali kutoka {ws.title}",
                        )
                    )

    db.session.commit()
    return f"imported {imported_days} day(s) across {len(month_sheets)} sheet(s)"
