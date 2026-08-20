from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

HEADER_FILL = PatternFill("solid", fgColor="D9E1F2")
HEADER_FONT = Font(name="Arial", bold=True, size=10)
BOLD_FONT = Font(name="Arial", bold=True, size=10)
BODY_FONT = Font(name="Arial", size=10)


def _sheet(wb, title, headers):
    ws = wb.active
    ws.title = title
    for col, text in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return ws


def _autosize(ws, n_cols):
    for col in range(1, n_cols + 1):
        letter = ws.cell(row=1, column=col).column_letter
        width = max(len(str(ws.cell(row=r, column=col).value or "")) for r in range(1, ws.max_row + 1))
        ws.column_dimensions[letter].width = max(12, min(40, width + 2))


def build_summary_excel(rows, totals, start, end):
    wb = Workbook()
    ws = _sheet(wb, "MUHTASARI", ["GARI", "MAKUSANYO", "MATUMIZI", "BAKI"])
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["car"]).font = BODY_FONT
        ws.cell(row=i, column=2, value=r["collected"]).font = BODY_FONT
        ws.cell(row=i, column=3, value=r["consumed"]).font = BODY_FONT
        ws.cell(row=i, column=4, value=r["net"]).font = BODY_FONT
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=1, value="JUMLA KUU").font = BOLD_FONT
    ws.cell(row=total_row, column=2, value=totals["grand_collected"]).font = BOLD_FONT
    ws.cell(row=total_row, column=3, value=totals["grand_consumed"]).font = BOLD_FONT
    ws.cell(row=total_row, column=4, value=totals["grand_net"]).font = BOLD_FONT
    ws.insert_rows(1)
    ws.cell(row=1, column=1, value=f"MUHTASARI: {start.isoformat()} hadi {end.isoformat()}").font = BOLD_FONT
    _autosize(ws, 4)
    return _to_bytes(wb)


def build_collections_excel(rows, total, start, end):
    wb = Workbook()
    ws = _sheet(
        wb, "MAKUSANYO",
        ["TAREHE YA MAKUSANYO", "TAREHE YA MUAMALA (BENKI)", "TRANS NO", "GARI", "KIASI", "MAELEZO"],
    )
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["date"].isoformat()).font = BODY_FONT
        ws.cell(row=i, column=2, value=r["trans_date"].isoformat()).font = BODY_FONT
        ws.cell(row=i, column=3, value=r["trans_no"]).font = BODY_FONT
        ws.cell(row=i, column=4, value=r["car"]).font = BODY_FONT
        ws.cell(row=i, column=5, value=r["amount"]).font = BODY_FONT
        ws.cell(row=i, column=6, value=r["note"]).font = BODY_FONT
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=4, value="JUMLA").font = BOLD_FONT
    ws.cell(row=total_row, column=5, value=total).font = BOLD_FONT
    _autosize(ws, 6)
    return _to_bytes(wb)


def build_consumption_excel(rows, total, start, end):
    wb = Workbook()
    ws = _sheet(wb, "MATUMIZI", ["TAREHE", "GARI", "AINA", "KIASI", "MAELEZO"])
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["date"].isoformat()).font = BODY_FONT
        ws.cell(row=i, column=2, value=r["car"]).font = BODY_FONT
        ws.cell(row=i, column=3, value=r["category"]).font = BODY_FONT
        ws.cell(row=i, column=4, value=r["amount"]).font = BODY_FONT
        ws.cell(row=i, column=5, value=r["description"]).font = BODY_FONT
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=3, value="JUMLA").font = BOLD_FONT
    ws.cell(row=total_row, column=4, value=total).font = BOLD_FONT
    _autosize(ws, 5)
    return _to_bytes(wb)


def build_reconciliation_excel(rows, total, start, end):
    wb = Workbook()
    ws = _sheet(wb, "UPATANISHO", ["TAREHE YA MUAMALA", "TRANS NO", "JUMLA", "MAELEZO"])
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["transaction_date"].isoformat()).font = BODY_FONT
        ws.cell(row=i, column=2, value=r["trans_no"]).font = BODY_FONT
        ws.cell(row=i, column=3, value=r["total"]).font = BODY_FONT
        ws.cell(row=i, column=4, value=r["note"]).font = BODY_FONT
    total_row = len(rows) + 2
    ws.cell(row=total_row, column=2, value="JUMLA KUU").font = BOLD_FONT
    ws.cell(row=total_row, column=3, value=total).font = BOLD_FONT
    _autosize(ws, 4)
    return _to_bytes(wb)


def build_shortfalls_excel(rows, start, end):
    wb = Workbook()
    ws = _sheet(wb, "UPUNGUFU", ["TAREHE", "GARI", "LENGO", "KILICHOKUSANYWA", "UPUNGUFU", "HALI", "MAELEZO"])
    for i, r in enumerate(rows, start=2):
        ws.cell(row=i, column=1, value=r["date"].isoformat()).font = BODY_FONT
        ws.cell(row=i, column=2, value=r["car"].code).font = BODY_FONT
        ws.cell(row=i, column=3, value=r["target"]).font = BODY_FONT
        ws.cell(row=i, column=4, value=r["collected"]).font = BODY_FONT
        ws.cell(row=i, column=5, value=r["shortfall"]).font = BODY_FONT
        ws.cell(row=i, column=6, value="Imefafanuliwa" if r["clearance"] else "Wazi").font = BODY_FONT
        ws.cell(row=i, column=7, value=r["clearance"].description if r["clearance"] else "").font = BODY_FONT
    _autosize(ws, 7)
    return _to_bytes(wb)


def _to_bytes(wb):
    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
